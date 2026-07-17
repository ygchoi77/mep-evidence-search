#!/usr/bin/env python3
"""Build a plaintext search index locally from the manual and audit workbook.

The generated JSON is intentionally gitignored. Run encrypt-vault.mjs before
publishing so GitHub only receives ciphertext.
"""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WEB_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = WEB_ROOT.parent
DATA_DIR = WEB_ROOT / "data"
TEXT_CACHE = PROJECT_ROOT / "tmp" / "pdfs" / "mechanical-manual.txt"
WORKBOOK = PROJECT_ROOT / "work" / "legal_audit" / "hierarchy-updated.xlsx"
MAX_MARKDOWN_LINE = 24285


def nfc(value: Any) -> str:
    return unicodedata.normalize("NFC", "" if value is None else str(value))


def compact(value: Any, limit: int | None = None) -> str:
    text = re.sub(r"\s+", " ", nfc(value)).strip()
    if limit and len(text) > limit:
        return text[: limit - 1].rstrip() + "…"
    return text


def match_text(value: Any) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", nfc(value)).lower()


def formula_url(value: Any) -> str:
    text = nfc(value)
    match = re.search(r'^=HYPERLINK\("([^"]+)"', text, re.IGNORECASE)
    return match.group(1).replace('""', '"') if match else ""


def locate_source(suffix: str) -> Path:
    candidates = [p for p in PROJECT_ROOT.glob(f"*{suffix}") if "기계설비" in nfc(p.name)]
    if not candidates:
        raise FileNotFoundError(f"기계설비 원본 {suffix} 파일을 찾지 못했습니다.")
    return candidates[0]


def extract_pdf_text(pdf_path: Path) -> list[str]:
    TEXT_CACHE.parent.mkdir(parents=True, exist_ok=True)
    if not TEXT_CACHE.exists() or TEXT_CACHE.stat().st_mtime < pdf_path.stat().st_mtime:
        executable = shutil.which("pdftotext")
        bundled = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/native/poppler/poppler/bin/pdftotext"
        if not executable and bundled.exists():
            executable = str(bundled)
        if executable:
            subprocess.run(
                [executable, "-layout", "-enc", "UTF-8", str(pdf_path), str(TEXT_CACHE)],
                check=True,
            )
        else:
            import pdfplumber

            with pdfplumber.open(pdf_path) as pdf:
                text = "\f".join(page.extract_text(layout=True) or "" for page in pdf.pages) + "\f"
            TEXT_CACHE.write_text(text, encoding="utf-8")

    pages = TEXT_CACHE.read_text(encoding="utf-8", errors="replace").split("\f")
    if pages and not pages[-1].strip():
        pages.pop()
    return pages


def page_title(raw: str, page: int, previous: str) -> str:
    lines = [compact(line) for line in raw.splitlines() if compact(line)]
    excluded = {"기계설비 기술기준 매뉴얼", "최 종 본", "목 차", "알 아 두 기 !"}
    numbered = re.compile(r"^(?:제\s*)?\d+(?:\.\d+){0,4}\.?\s+.{2,70}$")
    for line in lines[:24]:
        if line not in excluded and numbered.match(line) and not re.fullmatch(r"\d+", line):
            return line
    for line in lines[:14]:
        if line not in excluded and 4 <= len(line) <= 54 and not re.fullmatch(r"[\d.·\s]+", line):
            return line
    return previous or f"매뉴얼 {page}쪽"


def sheet_records(workbook, name: str):
    sheet = workbook[name]
    rows = sheet.iter_rows(values_only=True)
    headers = [compact(value) for value in next(rows)]
    for values in rows:
        if not any(value is not None and value != "" for value in values):
            continue
        yield {headers[index]: values[index] for index in range(min(len(headers), len(values)))}


def map_page(record: dict[str, Any], page_norms: list[str]) -> tuple[int | None, str]:
    try:
        line = int(record.get("줄 번호") or 0)
    except (TypeError, ValueError):
        line = 0
    estimated = max(1, min(len(page_norms), round((line / MAX_MARKDOWN_LINE) * len(page_norms)))) if line else 1

    values = [
        record.get("문맥 발췌"),
        record.get("원문 인용"),
        f"{record.get('원문 법령명') or ''}{record.get('원문 조문') or ''}",
        f"{record.get('정규 법령명') or ''}{record.get('정규 조문') or ''}",
    ]
    needles: list[str] = []
    for value in values:
        raw = nfc(value)
        for part in re.split(r"[…]|\.{3,}|\n", raw):
            needle = match_text(part)
            if len(needle) >= 12:
                needles.append(needle)
        whole = match_text(raw)
        if len(whole) >= 12:
            needles.append(whole)

    for needle in sorted(set(needles), key=len, reverse=True):
        matches = [index + 1 for index, page_text in enumerate(page_norms) if needle in page_text]
        if matches:
            return min(matches, key=lambda page: abs(page - estimated)), "exact"
    return None, "unmapped"


def main() -> None:
    pdf_path = locate_source(".pdf")
    page_raw = extract_pdf_text(pdf_path)
    page_norms = [match_text(page) for page in page_raw]
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=False)

    items: list[dict[str, Any]] = []
    previous_title = "기계설비 기술기준 매뉴얼"
    for index, raw in enumerate(page_raw, start=1):
        title = page_title(raw, index, previous_title)
        previous_title = title
        items.append(
            {
                "id": f"manual-page-{index}",
                "kind": "manual",
                "title": title,
                "source": "기계설비 기술기준 매뉴얼",
                "page": index,
                "section": title,
                "text": compact(raw, 9000),
                "status": "source",
                "date": "2022-05",
            }
        )

    mapped_citations = 0
    for row in sheet_records(workbook, "인용위치"):
        page, mapping = map_page(row, page_norms)
        mapped_citations += int(page is not None)
        confidence = compact(row.get("연결 신뢰도"))
        review_reason = compact(row.get("검토 사유"), 700)
        items.append(
            {
                "id": compact(row.get("인용 ID")) or f"citation-{len(items)}",
                "kind": "citation",
                "title": compact(row.get("조문·별표·별지/기준 항목 제목"))
                or compact(row.get("정규 조문"))
                or compact(row.get("정규 법령명")),
                "source": compact(row.get("정규 법령명")),
                "instrument": compact(row.get("정규 법령명")),
                "article": compact(row.get("정규 조문")),
                "hierarchy": compact(row.get("위계")),
                "section": compact(row.get("주변 제목") or row.get("HWPForge 섹션")),
                "page": page,
                "pageMapping": mapping,
                "line": row.get("줄 번호"),
                "quote": compact(row.get("원문 인용"), 1600),
                "context": compact(row.get("문맥 발췌"), 2400),
                "confidence": confidence,
                "reviewReason": review_reason,
                "officialUrl": formula_url(row.get("공식 원문 바로가기")),
                "status": "review" if review_reason or "낮" in confidence else "current",
            }
        )

    for row in sheet_records(workbook, "조문최신화"):
        change_status = compact(row.get("변경 상태"))
        items.append(
            {
                "id": compact(row.get("조문 ID")) or f"article-{len(items)}",
                "kind": "article",
                "title": compact(row.get("현행 조문/기준 항목 제목"))
                or compact(row.get("정규 조문")),
                "source": compact(row.get("정규 법령명")),
                "instrument": compact(row.get("정규 법령명")),
                "article": compact(row.get("정규 조문")),
                "hierarchy": compact(row.get("위계")),
                "section": compact(row.get("체계 경로")),
                "quote": compact(row.get("원문 인용 내용"), 1800),
                "text": compact(row.get("현행 조문/기준 항목 전문"), 7000),
                "summary": compact(row.get("변경 요약"), 1200),
                "changeStatus": change_status,
                "effectiveDate": compact(row.get("기준 시행일")),
                "officialUrl": formula_url(row.get("공식 URL")),
                "status": "review" if any(word in change_status for word in ("변경", "삭제", "폐지", "검토")) else "current",
            }
        )

    for row in sheet_records(workbook, "법령목록"):
        lookup_status = compact(row.get("조회 상태"))
        items.append(
            {
                "id": compact(row.get("법령 ID")) or f"instrument-{len(items)}",
                "kind": "instrument",
                "title": compact(row.get("정규 법령명")),
                "source": compact(row.get("조회 출처")),
                "instrument": compact(row.get("정규 법령명")),
                "hierarchy": compact(row.get("위계")),
                "section": compact(row.get("체계 경로")),
                "text": " ".join(
                    part
                    for part in [
                        compact(row.get("원문 표기")),
                        compact(row.get("세부 유형")),
                        compact(row.get("비고"), 1000),
                    ]
                    if part
                ),
                "effectiveDate": compact(row.get("시행일") or row.get("KCSC 갱신일")),
                "code": compact(row.get("KCSC 코드")),
                "officialUrl": formula_url(row.get("공식 URL")),
                "status": "review" if any(word in lookup_status for word in ("미확인", "후보", "오류", "검토")) else "current",
            }
        )

    report = json.loads((PROJECT_ROOT / "work" / "legal_audit" / "audit-report.json").read_text(encoding="utf-8"))
    payload = {
        "meta": {
            "title": "설비 근거검색",
            "source": "기계설비 기술기준 매뉴얼",
            "sourceDate": "2022-05",
            "snapshotDate": report.get("snapshotDate", "2026-07-17"),
            "pages": len(page_raw),
            "citations": report.get("occurrenceCount", 0),
            "articles": report.get("articleCount", 0),
            "instruments": report.get("instrumentCount", 0),
            "reviews": report.get("reviewCount", 0),
            "mappedCitations": mapped_citations,
            "notice": "설계 판단의 참고 자료이며 최종 적용 전 공식 원문과 프로젝트 조건을 확인해야 합니다.",
        },
        "items": items,
    }
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    output = DATA_DIR / "search-index.json"
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output),
                "items": len(items),
                "pages": len(page_raw),
                "mappedCitations": mapped_citations,
                "citations": report.get("occurrenceCount", 0),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
